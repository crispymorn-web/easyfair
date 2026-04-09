"""
openpyxl 기반 전문가 수준 견적서 Excel 생성
"""
import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

BRAND      = "2E7D4F"
BRAND_DARK = "1B5E20"
BRAND_LIGHT= "E8F5E9"
WHITE      = "FFFFFF"
GRAY_LIGHT = "F9FBF9"

SECTION_COLORS = {
    "SPACE_VENUE":    "1B5E20",
    "STRUCTURE":      "0D47A1",
    "GRAPHICS":       "1A5276",
    "AV_ELECTRICAL":  "4A148C",
    "FURNITURE":      "4E342E",
    "LOGISTICS":      "263238",
}
SECTION_BG = {
    "SPACE_VENUE":    "E8F5E9",
    "STRUCTURE":      "E3F2FD",
    "GRAPHICS":       "EBF5FB",
    "AV_ELECTRICAL":  "F3E5F5",
    "FURNITURE":      "FBE9E7",
    "LOGISTICS":      "ECEFF1",
}
SECTION_LABELS = {
    "SPACE_VENUE":    "1. Space & Venue Fees",
    "STRUCTURE":      "2. Structure & Construction",
    "GRAPHICS":       "3. Graphics & Signage",
    "AV_ELECTRICAL":  "4. AV & Electrical",
    "FURNITURE":      "5. Furniture & Accessories",
    "LOGISTICS":      "6. Logistics & Project Management",
}

def _fl(hex_c: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_c)

def _bd(style="thin", color="CCCCCC") -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _al(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

FMT_USD = '"$"#,##0'
FMT_KRW = '"₩"#,##0'


def generate_excel(quote: dict, items: list[dict]) -> bytes:
    """견적 데이터를 받아 Excel 바이트 반환"""
    wb = Workbook()
    ws = wb.active
    ws.title = "견적서"
    ws.sheet_view.showGridLines = False

    # 컬럼 너비
    for col, w in zip("ABCDEFGH", [5, 40, 8, 10, 14, 16, 18, 30]):
        ws.column_dimensions[get_column_letter(ord(col) - 64)].width = w

    usd_krw = quote.get("exchange_rate_usd_krw", 1370)

    # ── TITLE ──
    r = 1
    ws.row_dimensions[r].height = 28
    ws.merge_cells(f"A{r}:H{r}")
    c = ws[f"A{r}"]
    c.value = "EXHIBITION BOOTH CONSTRUCTION QUOTATION"
    c.font = Font(name="Arial", size=16, bold=True, color=WHITE)
    c.fill = _fl(BRAND_DARK)
    c.alignment = _al("center")

    r = 2
    ws.row_dimensions[r].height = 18
    ws.merge_cells(f"A{r}:H{r}")
    c = ws[f"A{r}"]
    c.value = f"{quote['event_name']}  |  {quote['venue_name']}, {quote['city']}"
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.fill = _fl(BRAND)
    c.alignment = _al("center")

    # 프로젝트 정보
    info = [
        ("Client",  quote.get("client_name", "")),
        ("Event",   quote.get("event_name", "")),
        ("Venue",   f"{quote.get('venue_name','')} ({quote.get('city','')}, {quote.get('country','')})"),
        ("Booth",   f"{quote.get('booth_width',0)}m × {quote.get('booth_depth',0)}m = {quote.get('booth_sqm',0)} sqm"),
        ("Date",    date.today().strftime("%B %d, %Y")),
        ("Rate",    f"1 USD = ₩{int(usd_krw):,}  KRW"),
    ]
    for k, v in info:
        r += 1
        ws.row_dimensions[r].height = 16
        ws.merge_cells(f"B{r}:H{r}")
        for cell, val, bold, bg, color in [
            (f"A{r}", k, True, BRAND, WHITE),
            (f"B{r}", v, False, "F1F8F4", "1A1A1A"),
        ]:
            c = ws[cell]
            c.value = val
            c.font = Font(name="Arial", size=10, bold=bold, color=color)
            c.fill = _fl(bg)
            c.alignment = _al("center" if bold else "left")
            c.border = _bd("thin", "AAAAAA")

    # 공백
    r += 1
    ws.row_dimensions[r].height = 7

    # ── TABLE HEADER ──
    r += 1
    ws.row_dimensions[r].height = 30
    for ci, h in enumerate(["No.", "Description", "Qty", "Unit",
                              "Unit Price (USD)", "Amount (USD)", "Amount (KRW)", "Remarks"], 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = _fl(BRAND)
        c.alignment = _al("center", wrap=True)
        c.border = _bd("medium", BRAND)

    # ── ITEMS BY SECTION ──
    section_order = [
        "SPACE_VENUE","STRUCTURE","GRAPHICS",
        "AV_ELECTRICAL","FURNITURE","LOGISTICS"
    ]
    subtotal_rows: list[int] = []

    for sec in section_order:
        sec_items = [i for i in items if i["section"] == sec]
        if not sec_items:
            continue

        # 섹션 헤더
        r += 1
        ws.row_dimensions[r].height = 20
        ws.merge_cells(f"A{r}:H{r}")
        c = ws[f"A{r}"]
        c.value = SECTION_LABELS.get(sec, sec)
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.fill = _fl(SECTION_COLORS.get(sec, BRAND))
        c.alignment = _al("left")
        c.border = _bd("medium", SECTION_COLORS.get(sec, BRAND))

        item_rows: list[int] = []
        for idx, item in enumerate(sec_items):
            r += 1
            ws.row_dimensions[r].height = 18
            alt = WHITE if idx % 2 == 0 else "FAFAFA"

            row_vals = [
                item.get("no", ""),
                item.get("description", ""),
                item.get("quantity", 0),
                item.get("unit", ""),
                item.get("unit_price_usd", 0),
                f"=C{r}*E{r}",
                f"=F{r}*{usd_krw}",
                item.get("notes", ""),
            ]
            aligns = ["center","left","center","center","right","right","right","left"]
            for ci, (val, aln) in enumerate(zip(row_vals, aligns), 1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font = Font(name="Arial", size=10, color="222222")
                c.fill = _fl(alt)
                c.alignment = _al(aln, wrap=(ci == 2))
                c.border = _bd()
                if ci == 5:
                    c.number_format = FMT_USD
                elif ci == 6:
                    c.number_format = FMT_USD
                elif ci == 7:
                    c.number_format = FMT_KRW

            item_rows.append(r)

        # 소계
        r += 1
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"A{r}:E{r}")
        c = ws[f"A{r}"]
        c.value = f"Subtotal — {SECTION_LABELS.get(sec, sec)}"
        c.font = Font(name="Arial", size=10, bold=True, color=SECTION_COLORS.get(sec, BRAND))
        c.fill = _fl(SECTION_BG.get(sec, BRAND_LIGHT))
        c.alignment = _al("right")
        c.border = _bd("medium", SECTION_COLORS.get(sec, BRAND))

        for col_l, fmt in [("F", FMT_USD), ("G", FMT_KRW)]:
            rng = f"{col_l}{item_rows[0]}:{col_l}{item_rows[-1]}"
            c = ws[f"{col_l}{r}"]
            c.value = f"=SUM({rng})"
            c.font = Font(name="Arial", size=10, bold=True, color=SECTION_COLORS.get(sec, BRAND))
            c.fill = _fl(SECTION_BG.get(sec, BRAND_LIGHT))
            c.alignment = _al("right")
            c.border = _bd("medium", SECTION_COLORS.get(sec, BRAND))
            c.number_format = fmt

        ws[f"H{r}"].fill = _fl(SECTION_BG.get(sec, BRAND_LIGHT))
        ws[f"H{r}"].border = _bd("medium", SECTION_COLORS.get(sec, BRAND))
        subtotal_rows.append(r)

    # ── GRAND TOTAL ──
    r += 1
    ws.row_dimensions[r].height = 26
    gt_f = "+".join(f"F{x}" for x in subtotal_rows)
    gt_g = "+".join(f"G{x}" for x in subtotal_rows)

    ws.merge_cells(f"A{r}:E{r}")
    c = ws[f"A{r}"]
    c.value = "GRAND TOTAL  (excl. applicable taxes)"
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.fill = _fl(BRAND)
    c.alignment = _al("right")
    c.border = _bd("medium", BRAND)

    for col_l, formula, fmt in [("F", f"={gt_f}", FMT_USD), ("G", f"={gt_g}", FMT_KRW)]:
        c = ws[f"{col_l}{r}"]
        c.value = formula
        c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
        c.fill = _fl(BRAND)
        c.alignment = _al("right")
        c.border = _bd("medium", BRAND)
        c.number_format = fmt

    ws[f"H{r}"].value = f"Rate: 1 USD = ₩{int(usd_krw):,}"
    ws[f"H{r}"].font = Font(name="Arial", size=9, color="A5D6A7")
    ws[f"H{r}"].fill = _fl(BRAND)
    ws[f"H{r}"].border = _bd("medium", BRAND)

    # 인쇄 설정
    ws.freeze_panes = "A11"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
